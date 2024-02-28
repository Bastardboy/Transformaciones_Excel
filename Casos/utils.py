import re
import os
import pandas as pd
import time
import csv
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import concurrent.futures
import requests

# Función para separar los números de las OV por un regex
def separar_numeros(s, regex_pattern):
    caracteres = [re.escape(c.strip()) for c in regex_pattern.split(',') if c.strip()]

    patron = '|'.join(caracteres)

    # Actualizado para tratar "rev. " seguido de un número como un solo valor
    regex_pattern = rf'(?i)(\d+)(?:{patron}(?:\d+|rev\. [12]\s*\d+))*'  # Actualizado para manejar "rev. 1" y "rev. 2"

    valores = re.findall(regex_pattern, s, flags=re.IGNORECASE)

    return [valor for valor in valores if not re.match(r'^\d$', valor)]

# Función para escribir la columna OV Limpio
def limpiar_ov(valor):
    if str(valor).isdigit():
        return int(valor)
    else:
        return valor


# SECCIÓN DE FUNCIONES PARA EL CASO DE LIMPIAR ARCHIVOS EN BASE A UNA CADENA REGEX

def procesar_multiples_archivos(archivos, regex_pattern, columns, columna_seleccionada, user_id, columna_orden=None):

    output_paths = []
    for archivo in archivos:
        columna_sel = columna_seleccionada.get(archivo)
        print(f'Columna seleccionada: {columna_sel} del archivo {archivo}')
        if columna_sel is not None:
            output_path = procesar_archivo(archivo, regex_pattern, user_id, columns, columna_sel, columna_orden)
            print(f"Ruta de salida: {output_path}")  # Debug: Verificar la ruta de salida
            output_paths.append(output_path)
        # Eliminar todos los archivos procesados
    for archivo in archivos:
        if os.path.exists(archivo):
            os.remove(archivo)
            
    print(f'Output paths: {output_paths}')
    return output_paths

def procesar_archivo(archivo, regex_pattern, user_id, columnas, columna_separacion, column_order=None):

    ruta_raiz = 'Limpiados'
    ruta_usuario = os.path.join(ruta_raiz, user_id)
    
    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)

    df = pd.read_excel(archivo, header=0)

    nuevas_filas = []
    # Tiempo de Procesamiento
    for col in columnas:
        if col not in df.columns:
            df[col] = '0'

    for _, row in df.iterrows():
        valores_separados = separar_numeros(str(row.get(columna_separacion, '')), regex_pattern)

        if not valores_separados:
            nueva_fila = row.copy()
            nueva_fila[columna_separacion + ' LIMPIO'] = row.get(columna_separacion, '')
            nuevas_filas.append(nueva_fila)
        else:
            for i, valor in enumerate(valores_separados):
                nueva_fila = row.copy()
                nueva_fila[columna_separacion] = valor
                nueva_fila[columna_separacion + ' LIMPIO'] = limpiar_ov(valor)

                if i != 0:
                    for col in columnas:
                        nueva_fila[col] = ''
                
                nuevas_filas.append(nueva_fila)

    df_extendido = pd.DataFrame(nuevas_filas)

    if column_order:
        df_extendido = df_extendido[column_order]
    
    # Guardar el DataFrame en un archivo Excel
    output_path = os.path.join(ruta_usuario, f"procesado_{os.path.basename(archivo)}")
    df_extendido.to_excel(output_path, index=False)

    return output_path


def cargar_archivos_limpiar(archivos_base, user_id):
    ruta_temporal = 'Limpiados'
    ruta_usuario = os.path.join(ruta_temporal, user_id)

    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)

    headers = []  # Lista para almacenar los encabezados del Archivo Base
    saved_files_base = []

    for archivo_base in archivos_base:
        try:
            file_path = archivo_base
            saved_files_base.append(file_path)
            # Leer solo los encabezados del archivo con openpyxl
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            headers.append([cell.value for cell in ws[1]])

        except FileNotFoundError as e:
            print(f'Error al abrir el archivo base: {e}')
        except PermissionError as e:
            print(f'Error con respecto a los permisos: {e}')
        except Exception as e:
            print(f'Error al procesar el archivo {archivo_base}: {e}')


    return saved_files_base, headers


# SECCIÓN DE FUNCIONES PARA EL CASO DE LIMPIAR ARCHIVOS DETALLADOS, CSV EN BASE A TUBO O COMAS.

def limpiar_detalle(archivo, modo_seleccionado, user_id):
    ruta_temporal = 'Detallados'
    ruta_usuario = os.path.join(ruta_temporal, user_id)

    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)

    data = []
    max_fields = 0
    try:
        # Determinar el delimitador basado en el modo seleccionado
        delimiter = ',' if modo_seleccionado == 'limpiar_coma' else '|'

        with open(archivo, 'r', encoding='latin1') as file:
            reader = csv.reader(file, delimiter=delimiter)
            for row in reader:
                data.append(row)
                if len(row) > max_fields:
                    max_fields = len(row)

        # Rellenar las filas con menos campos con NaN
        for i in range(len(data)):
            while len(data[i]) < max_fields:
                data[i].append(None)

        # Convertir los datos a un DataFrame
        if modo_seleccionado == 'limpiar_coma':
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(data)

        try: 
            df.iloc[:, 15:] = df.iloc[:, 15:].apply(pd.to_numeric, errors='ignore')
            # Redondear los valores de la columna J (índice 9) a 1 decimal
            df.iloc[:, 9] = df.iloc[:, 9].round(1)

            # Convertir los valores de la columna I (índice 8) a enteros
            df.iloc[:, 8] = df.iloc[:, 8].astype(int)
        except Exception as e:
            print(f'Error al convertir a números: {e}')
        
        nombre_archivo = os.path.basename(archivo)
        try: 
            # Guardar el DataFrame en un archivo Excel
            nombre_base, extension = os.path.splitext(nombre_archivo)
            output_file_path = os.path.join(ruta_usuario, f'Archivo_{nombre_base}_Limpiado.xlsx')
            df.to_excel(output_file_path, index=False, header=None)

           
            return output_file_path
        except Exception as e:
            print(f'Error al guardar el archivo: {e}')
            return None

    except Exception as e:
        print(f'Error al leer el archivo: {e}')
        return None

# SECCIÓN DE FUNCIONES PARA EL CASO DE CONSOLIDAR ARCHIVOS, TANTO PARA EL CASO GLOBAL, COMO PARA UN CASO ESPECÍFICO

def cargar_archivo(archivos, user_id):
    print(f'Archivos recibidos: {archivos}')
    ruta_temporal = 'Consolidados'
    ruta_usuario = os.path.join(ruta_temporal, user_id)

    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)

    df_final = None
    archivos_temporales = []

    try:
        for archivo in archivos:
            try:
                file_path = os.path.join(ruta_usuario, archivo.filename)
                archivo.save(file_path)
                archivos_temporales.append(file_path)

                # Lee el DataFrame desde el archivo y agrega a la lista
                df_actual = pd.read_excel(file_path, header=0)

                # Concatena todos los DataFrames en uno solo
                df_final = pd.concat([df_final, df_actual])
            except Exception as e:
                print(f'Error al procesar el archivo {archivo.filename}: {e}')

        nombre_archivo_final = f'Archivo_Consolidado.xlsx'
        ruta_archivo_final = os.path.join(ruta_usuario, nombre_archivo_final)
        print(f'Ruta del archivo final: {ruta_archivo_final}')
        df_final.to_excel(ruta_archivo_final, index=False)

        # Borrar archivos temporales
        for file_path in archivos_temporales:
            os.remove(file_path)

        return ruta_archivo_final

    except Exception as e:
        print(f'Error al guardar el archivo consolidado: {e}')
        return None


def cargar_archivos(archivos_base, archivos_combinar, user_id):
    ruta_temporal = 'Consolidados'
    ruta_usuario = os.path.join(ruta_temporal, user_id)

    print(f'archivos_base: {archivos_base}')
    print(f'archivos_combinar: {archivos_combinar}')

    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)

    headers_base = None  # Lista para almacenar los encabezados del Archivo Base
    headers_combinar = None  # Lista para almacenar los encabezados del Archivo a Combinar

    saved_files_base = []
    saved_files_combinar = []

    for archivo_base in archivos_base:
        try:
            file_path = os.path.join(ruta_usuario, archivo_base.filename)
            archivo_base.save(file_path)
            saved_files_base.append(file_path)

            # Leer solo los encabezados del archivo con openpyxl
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            headers_base = [cell.value for cell in ws[1]]

        except Exception as e:
            print(f'Error al procesar el archivo {archivo_base.filename}: {e}')

    for archivo_combinar in archivos_combinar:
        try:
            file_path = os.path.join(ruta_usuario, archivo_combinar.filename)
            archivo_combinar.save(file_path)
            saved_files_combinar.append(file_path)

            # Leer solo los encabezados del archivo con openpyxl
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            headers_combinar = [cell.value for cell in ws[1]]

        except Exception as e:
            print(f'Error al procesar el archivo {archivo_combinar.filename}: {e}')

    return saved_files_base, headers_base, saved_files_combinar, headers_combinar

def consolidar_archivos(archivo_base, archivos_combinar, header_base, header_combinar, columnas_seleccionadas, user_id):
    ruta_temporal = 'Consolidados'
    ruta_usuario = os.path.join(ruta_temporal, user_id)

    if not os.path.exists(ruta_usuario):
        os.makedirs(ruta_usuario)
    print(f'Columnas seleccionadas en la función consolidar archivo, antes del try y for: {columnas_seleccionadas}')
    try:
        archivo_base = os.path.normpath(archivo_base)
        archivos_combinar = [os.path.normpath(archivo) for archivo in archivos_combinar]
        df_base = pd.read_excel(archivo_base, header=0)
        
        # Guardar los nombres de las columnas del archivo base
        columnas_base = df_base.columns.tolist()

        print(f'Empieza a contar el tiempo')
        start_time = time.time()
        # Leer y combinar los archivos a fusionar
        for archivo_combinar in archivos_combinar:
            df_combinar = pd.read_excel(archivo_combinar, header=0)

            df_base = pd.merge(df_base, df_combinar, left_on=header_base, right_on=header_combinar, how='left')
        
        print(f'Columnas dentro del try{columnas_seleccionadas}')
        # Seleccionar las columnas del archivo base y las columnas seleccionadas de los archivos a combinar
        columnas_finales = columnas_base + columnas_seleccionadas
        df_base = df_base[columnas_finales]

        nombre_archivo_base = os.path.basename(archivo_base) 
        # Guardar el resultado en un nuevo archivo
        resultado_path = os.path.join(ruta_usuario,  f'Consolidación_para_{nombre_archivo_base}')
        df_base.to_excel(resultado_path, index=False)
        print(f'Ruta del archivo final: {resultado_path}')
        end_time = time.time()
        print(f'Tiempo de ejecución para el merge: {end_time - start_time} segundos')
        return resultado_path

    except Exception as e:
        print(f'Error al consolidar archivos: {e}')
        return None
    
def validar_pdf(ruta_archivo):
    try:
        with open(ruta_archivo, 'rb') as archivo_pdf:
            pdf_reader = PdfReader(archivo_pdf)
            return len(pdf_reader.pages) > 0
    except Exception as e:
        return False
    

def descargar_pdf(folio, tipo, rut, resolucion, nombre_empresa, directorio_usuario):
    enlace = f'http://soaoci2.grupogtd.com/DTEPlus/getFactura.jsp?folio={folio}&tipo={tipo}&rut={rut}&resolucion={resolucion}&=.pdf'
    nombre_archivo = f'{nombre_empresa}_{folio}.pdf'
    ruta_destino = os.path.join(directorio_usuario, nombre_archivo)

    try:
        with requests.get(enlace, stream=True) as respuesta:
            respuesta.raise_for_status()

            if "application/pdf" in respuesta.headers.get("content-type", ""):
                with open(ruta_destino, 'wb') as archivo_destino:
                    archivo_destino.write(respuesta.content)

                if validar_pdf(ruta_destino):
                    return {'nombre_empresa': nombre_empresa, 'folio': folio, 'enlace_descarga': f'/descargas/{nombre_archivo}'}
                else:
                    os.remove(ruta_destino)
                    return f"Error al validar el PDF para el folio {folio}"
            else:
                return f"El enlace {enlace} no devuelve un archivo PDF."
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            return f"El recurso no se encontró para el folio {folio}"
        else:
            return f"Error al descargar {enlace}: {e}"
    except Exception as e:
        return f"Error inesperado: {e}"
    
def descargar_en_paralelo(parametros, directorio_usuario):
    folios = parametros['numeros_de_folio']
    tipo = parametros['tipo']
    rut = parametros['rut']
    resolucion = parametros['resolucion']
    nombre_empresa = parametros['nombre_empresa']

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(descargar_pdf, folio, tipo, rut, resolucion, nombre_empresa, directorio_usuario) for folio in folios]

        enlaces_descarga = []
        errores = []
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            if isinstance(result, dict):
                enlaces_descarga.append(result)
            else:
                errores.append(result)

    return enlaces_descarga, errores

def eliminar_archivo_despues_de_tiempo(ruta_archivo, delay):
    time.sleep(delay)
    if os.path.exists(ruta_archivo):
        os.remove(ruta_archivo)